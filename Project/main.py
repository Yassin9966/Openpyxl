from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# creating workbook object
wb = Workbook()

# loading existing spredsheet
wb = load_workbook('openpyxl_project.xlsx')

# creating worksheet
ws = wb.active

# print out data
for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print()

# adding 10,11,12
'''ws['A5'] = 10
ws['B5'] = 11
ws['C5'] = 12
wb.save('openpyxl_project.xlsx')'''

# adding sheet
'''wb.create_sheet('Tabelle2')
wb.save('openpyxl_project.xlsx')'''

# accessing second sheet 'Tabelle2'
second_sheet = wb['Tabelle2']

# adding A, 10, 20, 30, 40, 50
'''second_sheet['A1'] = "A"
second_sheet['A2'] = 10
second_sheet['A3'] = 20
second_sheet['A4'] = 30
second_sheet['A5'] = 40
second_sheet['A6'] = 50
wb.save('openpyxl_project.xlsx')'''

print("\n--------------------------\n")

# print out row
row = second_sheet['A1':'A6']
for x in row:
    for y in x:
        print(y.value)

# check if even or odd (with the help of chatgbt)
for cell in second_sheet['A']:
    if isinstance(cell.value, (int, float)):
        result = 'Even' if cell.value % 2 == 0 else 'Odd'
        second_sheet.cell(row=cell.row, column=cell.column + 1).value = result

# saving
'''wb.save('openpyxl_project.xlsx')'''